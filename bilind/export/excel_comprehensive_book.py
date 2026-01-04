"""
Comprehensive Quantity Book Export
==================================
Refactored to strictly adhere to Single Source of Truth (UnifiedCalculator).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Any, Callable, Optional, List

# ... (Keep existing Style constants: COLOR_HEADER_BG, FONT_HEADER, etc.) ...
COLOR_HEADER_BG = "1F4E78"
COLOR_HEADER_FG = "FFFFFF"
COLOR_TOTAL_BG = "BDD7EE"
FONT_HEADER = Font(name='Segoe UI', bold=True, size=11, color=COLOR_HEADER_FG)
FONT_NORMAL = Font(name='Segoe UI', size=10)
FONT_BOLD = Font(name='Segoe UI', bold=True, size=10)
BORDER_THIN = Border(left=Side(style='thin',color='BFBFBF'), right=Side(style='thin',color='BFBFBF'), top=Side(style='thin',color='BFBFBF'), bottom=Side(style='thin',color='BFBFBF'))
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def setup_sheet(ws, title):
    ws.title = title
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

def write_header(ws, headers):
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=text)
        c.font = FONT_HEADER; c.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        c.alignment = ALIGN_CENTER; c.border = BORDER_THIN

def write_row(ws, row, data):
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_NORMAL; c.border = BORDER_THIN; c.alignment = ALIGN_CENTER
        if isinstance(val, (int, float)): c.number_format = '0.00'

def auto_fit(ws):
    for column in ws.columns:
        ws.column_dimensions[get_column_letter(column[0].column)].width = 20

def export_comprehensive_book(project: Any, filepath: str, app: Any = None, status_cb = None, selected_sheets = None) -> bool:
    if not filepath:
        # Backward-compatible: allow callers to omit filepath and prompt via app.
        if app and hasattr(app, '_ask_save_path'):
            filepath = app._ask_save_path('دفتر_الكميات_الشامل.xlsx', 'excel')
        if not filepath:
            return False

    if status_cb:
        status_cb("Generating Book...", "📊")
    
    # 1. INITIALIZE SSOT
    from ..calculations.unified_calculator import UnifiedCalculator
    calc = UnifiedCalculator(project)
    
    # Pre-calculate EVERYTHING
    project_totals = calc.calculate_totals()
    all_rooms_data = calc.calculate_all_rooms()
    rooms_map = {r.room_name: r for r in all_rooms_data} # Fast Lookup
    
    wb = openpyxl.Workbook()
    # We'll always generate our own sheets; remove default placeholder.
    try:
        wb.remove(wb.active)
    except Exception:
        pass
    
    def val(obj, attr, default=None):
        return obj.get(attr, default) if isinstance(obj, dict) else getattr(obj, attr, default)

    def fnum(x: Any) -> float:
        try:
            return float(x or 0.0)
        except Exception:
            return 0.0

    def get_opening_area(o: Any) -> float:
        # Prefer model's computed area if available.
        area = val(o, 'area', None)
        if area is not None:
            return fnum(area)
        w = fnum(val(o, 'width', None) or val(o, 'w', 0.0))
        h = fnum(val(o, 'height', None) or val(o, 'h', 0.0))
        qty = int(val(o, 'quantity', None) or val(o, 'qty', 1) or 1)
        return max(0.0, w * h * qty)

    def get_opening_stone(o: Any) -> float:
        stone = val(o, 'stone', None)
        if stone is not None:
            return fnum(stone)
        w = fnum(val(o, 'width', None) or val(o, 'w', 0.0))
        h = fnum(val(o, 'height', None) or val(o, 'h', 0.0))
        qty = int(val(o, 'quantity', None) or val(o, 'qty', 1) or 1)
        # Match UnifiedCalculator.calculate_stone() behavior.
        otype = str(val(o, 'opening_type', '') or '').upper()
        if otype == 'DOOR':
            perim_each = (2 * h) + w
        else:
            perim_each = 2 * (h + w)
        return max(0.0, perim_each * qty)

    # Detailed-books helpers (local to exporter)
    from ..calculations.helpers import safe_zone_area

    def norm_text(x: Any) -> str:
        try:
            return str(x or '').strip().lower()
        except Exception:
            return ''

    def opening_area_each(o: Any) -> float:
        # area_each if available, else width*height
        ae = val(o, 'area_each', None)
        if ae is not None:
            return fnum(ae)
        w = fnum(val(o, 'width', None) or val(o, 'w', 0.0))
        h = fnum(val(o, 'height', None) or val(o, 'h', 0.0))
        return max(0.0, w * h)

    def opening_rooms(o: Any) -> List[str]:
        # Combine assigned_rooms + room.opening_ids references
        name = str(val(o, 'name', '') or '').strip()
        assigned = val(o, 'assigned_rooms', None) or []
        seen = set()
        out: List[str] = []
        for rn in assigned:
            rname = str(rn or '').strip()
            if rname and rname not in seen:
                seen.add(rname)
                out.append(rname)
        if name:
            for r in getattr(project, 'rooms', []) or []:
                rname = str(val(r, 'name', '') or '').strip()
                ids = val(r, 'opening_ids', []) or []
                if name in ids and rname and rname not in seen:
                    seen.add(rname)
                    out.append(rname)
        return out

    def opening_area_for_room(o: Any, room_name: str) -> float:
        room_name = str(room_name or '').strip()
        if not room_name:
            return 0.0

        # 1) room_quantities (strongest)
        rq = val(o, 'room_quantities', None) or {}
        if rq and room_name in rq:
            qty = int(rq.get(room_name, 0) or 0)
            return opening_area_each(o) * max(0, qty)

        # 2) room_shares
        shares = val(o, 'room_shares', None) or {}
        if shares:
            return get_opening_area(o) * max(0.0, fnum(shares.get(room_name, 0.0)))

        # 3) assigned_rooms + share_mode
        assigned = val(o, 'assigned_rooms', None) or []
        mode = val(o, 'share_mode', None) or None
        if assigned:
            if room_name not in assigned:
                return 0.0
            if (mode or 'split') == 'single':
                return get_opening_area(o) if room_name == str(assigned[0] or '').strip() else 0.0
            return get_opening_area(o) / max(1, len(assigned))

        # 4) fallback split among rooms that reference it in opening_ids
        rooms_ref = opening_rooms(o)
        if rooms_ref and room_name in rooms_ref:
            return get_opening_area(o) / max(1, len(rooms_ref))

        return 0.0

    default_h = fnum(getattr(project, 'default_wall_height', 3.0) or 3.0)

    def iter_room_walls(r: Any) -> List[Any]:
        walls = val(r, 'walls', []) or []
        if walls:
            return list(walls)
        perim = fnum(val(r, 'perimeter', 0.0) or val(r, 'perim', 0.0))
        rh = fnum(val(r, 'wall_height', 0.0) or val(r, 'height', 0.0) or default_h)
        if perim > 0 and rh > 0:
            return [{
                'name': 'محيط الغرفة',
                'length': perim,
                'height': rh,
                '_pseudo': True,
            }]
        return []

    def wall_name(w: Any, idx: int) -> str:
        n = str(val(w, 'name', '') or '').strip()
        return n if n else f"Wall {idx + 1}"

    def wall_length(w: Any) -> float:
        return max(0.0, fnum(val(w, 'length', 0.0)))

    def wall_height(w: Any, r: Any) -> float:
        wh = fnum(val(w, 'height', 0.0))
        if wh > 0:
            return wh
        rh = fnum(val(r, 'wall_height', 0.0) or val(r, 'height', 0.0))
        return rh if rh > 0 else default_h

    def allocate_capped(total: float, weights: List[float], caps: List[float]) -> (List[float], float):
        """Allocate `total` across items proportionally by weights, capped by caps.

        Returns (allocations, unallocated).
        """
        total = fnum(total)
        caps_left = [max(0.0, fnum(c)) for c in caps]
        wts = [max(0.0, fnum(w)) for w in weights]
        alloc = [0.0 for _ in caps_left]
        remaining = total
        active = [i for i, c in enumerate(caps_left) if c > 0]
        eps = 1e-9

        while remaining > eps and active:
            wsum = sum(wts[i] for i in active)
            if wsum <= eps:
                wsum = float(len(active))
                for i in active:
                    wts[i] = 1.0
            exhausted = []
            for i in active:
                share = remaining * (wts[i] / wsum) if wsum > eps else 0.0
                if share <= caps_left[i] + eps:
                    alloc[i] += share
                    caps_left[i] -= share
                else:
                    alloc[i] += caps_left[i]
                    remaining -= caps_left[i]
                    caps_left[i] = 0.0
                    exhausted.append(i)
            # If we didn't exhaust anything, we're done
            if not exhausted:
                remaining = 0.0
                break
            active = [i for i in active if caps_left[i] > eps]

        unallocated = max(0.0, remaining)
        return alloc, unallocated

    # Build per-room opening objects (support both room.opening_ids and opening.assigned_rooms)
    all_openings = (getattr(project, 'doors', []) or []) + (getattr(project, 'windows', []) or [])
    openings_by_name = {str(val(o, 'name', '') or '').strip(): o for o in all_openings if str(val(o, 'name', '') or '').strip()}

    room_openings_map = {}
    for r in getattr(project, 'rooms', []) or []:
        rname = str(val(r, 'name', '') or '').strip()
        refs = []
        for oid in (val(r, 'opening_ids', []) or []):
            o = openings_by_name.get(str(oid or '').strip())
            if o is not None:
                refs.append(o)
        for o in all_openings:
            assigned = val(o, 'assigned_rooms', None) or []
            room_shares = val(o, 'room_shares', None) or {}
            room_qtys = val(o, 'room_quantities', None) or {}
            if rname and (rname in assigned or rname in room_shares or rname in room_qtys):
                refs.append(o)
        dedup = {}
        for o in refs:
            nm = str(val(o, 'name', '') or '').strip()
            if nm:
                dedup[nm] = o
        room_openings_map[rname] = list(dedup.values())

    def ensure_sheet(key: str) -> bool:
        return (not selected_sheets) or (key in (selected_sheets or []))

    def add_totals_row(ws, row: int, label: str, value: float, unit: str = ""):
        ws.cell(row=row, column=2, value=label).font = FONT_BOLD
        ws.cell(row=row, column=3, value=float(value or 0.0)).number_format = '0.00'
        if unit:
            ws.cell(row=row, column=4, value=unit)
    
    # --- SHEET: SUMMARY ---
    if ensure_sheet('summary'):
        ws = wb.create_sheet("ملخص المشروع")
        setup_sheet(ws, "ملخص المشروع")
        
        ws['B2'] = "ملخص الكميات (SSOT Verified)"; ws['B2'].font = FONT_HEADER
        
        doors = getattr(project, 'doors', []) or []
        windows = getattr(project, 'windows', []) or []
        doors_area = sum(get_opening_area(o) for o in doors)
        windows_area = sum(get_opening_area(o) for o in windows)
        openings_area = doors_area + windows_area
        openings_count = len(doors) + len(windows)

        tiles_items = getattr(project, 'tiles_items', []) or []
        tiles_total = sum(fnum(val(i, 'area', 0.0)) for i in tiles_items)

        baseboards_items = getattr(project, 'baseboards', []) or []
        baseboards_items_total = sum(fnum(val(b, 'net_length_m', None) or val(b, 'length', None) or 0.0) for b in baseboards_items)

        data = [
            ("عدد الغرف", len(getattr(project, 'rooms', []) or []), ""),
            ("مساحات الأرضيات (إجمالي)", project_totals.get('area_total', 0.0), "م²"),
            ("أعمال الزريقة (SSOT)", project_totals.get('plaster_total', 0.0), "م²"),
            ("أعمال الدهان (SSOT)", project_totals.get('paint_total', 0.0), "م²"),
            ("سيراميك الجدران (SSOT)", project_totals.get('ceramic_wall', 0.0), "م²"),
            ("سيراميك الأرضيات (SSOT)", project_totals.get('ceramic_floor', 0.0), "م²"),
            ("إجمالي السيراميك (SSOT)", project_totals.get('ceramic_total', 0.0), "م²"),
            ("نعلات (SSOT)", project_totals.get('baseboard_total', 0.0), "م.ط"),
            ("حجر/أطر (SSOT)", project_totals.get('stone_total', 0.0), "م.ط"),
            ("مساحة الفتحات (أبواب+شبابيك)", openings_area, "م²"),
            ("عدد الفتحات (أبواب+شبابيك)", openings_count, ""),
            ("بلاط الأرضيات (دفتر البنود)", tiles_total, "م²"),
            ("نعلات (دفتر البنود)", baseboards_items_total, "م.ط"),
        ]

        for i, (k, v, u) in enumerate(data, 5):
            add_totals_row(ws, i, k, v, u)
            
    # --- SHEET: ROOMS ---
    if ensure_sheet('rooms'):
        ws = wb.create_sheet("مساحة الغرف")
        setup_sheet(ws, "مساحة الغرف")
        write_header(ws, [
            "م", "الغرفة", "النوع", "المحيط", "ارتفاع الجدار",
            "مساحة الأرضية", "جدران قائمة", "خصم فتحات", "جدران صافية",
            "زريقة جدران", "زريقة سقف", "دهان جدران", "دهان سقف",
            "سيراميك جدران", "سيراميك أرضيات", "سيراميك سقف",
            "نعلات", "حجر/أطر"
        ])
        
        for i, r in enumerate(project.rooms, 2):
            name = val(r, 'name', '-')
            rtype = val(r, 'room_type', '')
            perim = fnum(val(r, 'perimeter', 0.0))
            # Wall height: check if room has multiple walls with different heights
            walls = val(r, 'walls', [])
            if walls and len(walls) > 1:
                heights = set()
                for w_item in walls:
                    wh_val = fnum(val(w_item, 'height', 0.0))
                    if wh_val <= 0:
                        wh_val = fnum(val(r, 'wall_height', 0.0) or val(r, 'height', 0.0))
                    if wh_val > 0:
                        heights.add(round(wh_val, 2))
                if len(heights) > 1:
                    min_h = min(heights)
                    max_h = max(heights)
                    wh = f"{min_h:.2f}-{max_h:.2f}"  # Show range as text
                elif heights:
                    wh = next(iter(heights))
                else:
                    wh = fnum(val(r, 'wall_height', 0.0) or val(r, 'height', 0.0))
            else:
                wh = fnum(val(r, 'wall_height', 0.0) or val(r, 'height', 0.0))
            # FETCH FROM SSOT
            d = rooms_map.get(name)
            
            if d:
                write_row(ws, i, [
                    i-1, name, rtype, perim, wh,
                    d.ceiling_area, d.walls_gross,
                    d.walls_openings, d.walls_net,
                    d.plaster_walls, d.plaster_ceiling,
                    d.paint_walls, d.paint_ceiling,
                    d.ceramic_wall, d.ceramic_floor, d.ceramic_ceiling,
                    d.baseboard_length, d.stone_length
                ])
            else:
                write_row(ws, i, [i-1, name, rtype, perim, wh] + [0] * 13)

        auto_fit(ws)

    # --- SHEET: PLASTER ---
    if ensure_sheet('plaster'):
        ws = wb.create_sheet("أعمال الزريقة")
        setup_sheet(ws, "أعمال الزريقة")
        write_header(ws, ["م", "الغرفة", "العنصر", "المساحة القائمة", "خصم", "الصافي"])
        
        r_idx = 2
        for r in project.rooms:
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            if not d: continue
            
            # Wall Row (Get exact values from calculator)
            # Net Plaster = d.plaster_walls
            # Gross = d.walls_gross
            # Deduct = Gross - Net
            w_deduct = d.walls_gross - d.plaster_walls
            write_row(ws, r_idx, [r_idx-1, name, "جدران", d.walls_gross, w_deduct, d.plaster_walls])
            r_idx += 1
            
            # Ceiling Row
            c_deduct = d.ceiling_area - d.plaster_ceiling
            write_row(ws, r_idx, [r_idx-1, name, "سقف", d.ceiling_area, c_deduct, d.plaster_ceiling])
            r_idx += 1

        auto_fit(ws)

        # Detailed per-wall plaster book (standalone)
        ws2 = wb.create_sheet("أعمال الزريقة (مفصل)")
        setup_sheet(ws2, "أعمال الزريقة (مفصل)")
        write_header(ws2, [
            "م", "الغرفة", "الجدار/السقف", "الطول (م)", "الارتفاع (م)",
            "قائم (م²)", "خصم فتحات (م²)", "خصم سيراميك (م²)", "صافي الزريقة (م²)", "ملاحظات"
        ])
        row = 2
        seq = 1
        for r in getattr(project, 'rooms', []) or []:
            rname = str(val(r, 'name', '') or '').strip()
            walls = iter_room_walls(r)
            if not walls:
                continue
            d = rooms_map.get(rname)
            ssot_open = fnum(getattr(d, 'walls_openings', 0.0) or 0.0) if d else 0.0
            ssot_plaster = fnum(getattr(d, 'plaster_walls', 0.0) or 0.0) if d else 0.0
            ssot_cer = fnum(getattr(d, 'ceramic_wall', 0.0) or 0.0) if d else 0.0
            plaster_under = bool(getattr(project, 'plaster_under_ceramic', True))

            openings = room_openings_map.get(rname, [])
            hosted = [o for o in openings if val(o, 'host_wall', None)]
            lengths = [wall_length(w) for w in walls]
            heights = [wall_height(w, r) for w in walls]
            names = [wall_name(w, i) for i, w in enumerate(walls)]
            gross_list = [max(0.0, lengths[i] * heights[i]) for i in range(len(walls))]

            # Hosted opening demand per wall
            hosted_dem = [0.0 for _ in walls]
            for o in hosted:
                host = norm_text(val(o, 'host_wall', ''))
                if not host:
                    continue
                for iwn, wn in enumerate(names):
                    if host == norm_text(wn):
                        hosted_dem[iwn] += opening_area_for_room(o, rname)
                        break

            total_hosted = sum(hosted_dem)
            if ssot_open <= 0 and total_hosted > 0:
                ssot_open = total_hosted
            scale = 1.0
            notes_open = ''
            if total_hosted > 0 and ssot_open > 0 and total_hosted > ssot_open:
                scale = ssot_open / total_hosted
                notes_open = 'تم تحجيم فتحات مستضافة لتطابق SSOT'
            hosted_alloc = [min(gross_list[i], hosted_dem[i] * scale) for i in range(len(walls))]
            rem_caps = [max(0.0, gross_list[i] - hosted_alloc[i]) for i in range(len(walls))]
            rem_open = max(0.0, ssot_open - sum(hosted_alloc))
            dist_open, unalloc_open = allocate_capped(rem_open, lengths, rem_caps)
            open_alloc = [hosted_alloc[i] + dist_open[i] for i in range(len(walls))]
            if unalloc_open > 0.01:
                notes_open = (notes_open + '؛ ' if notes_open else '') + 'فتحـات أكبر من طاقة الجدران'
            net_after_open = [max(0.0, gross_list[i] - open_alloc[i]) for i in range(len(walls))]

            # Ceramic allocation per wall for optional plaster deduction
            zones = getattr(project, 'ceramic_zones', []) or []
            room_z = [z for z in zones if norm_text(val(z, 'room_name', '')) == norm_text(rname)]
            wall_z = [z for z in room_z if norm_text(val(z, 'surface_type', 'wall')) == 'wall']
            z_hosted = [z for z in wall_z if val(z, 'wall_name', None)]
            hosted_cer = [0.0 for _ in walls]
            for z in z_hosted:
                wname = norm_text(val(z, 'wall_name', ''))
                if not wname:
                    continue
                for iwn, wn in enumerate(names):
                    if wname == norm_text(wn):
                        hosted_cer[iwn] += fnum(safe_zone_area(z))
                        break
            target_cer = min(ssot_cer, sum(net_after_open)) if ssot_cer > 0 else 0.0
            sum_hosted_cer = sum(hosted_cer)
            cer_alloc = [0.0 for _ in walls]
            notes_cer = ''
            if sum_hosted_cer > 0 and target_cer > 0 and sum_hosted_cer > target_cer:
                s = target_cer / sum_hosted_cer
                cer_alloc = [min(net_after_open[i], hosted_cer[i] * s) for i in range(len(walls))]
                notes_cer = 'تم تحجيم سيراميك مستضاف لتطابق SSOT'
            else:
                cer_alloc = [min(net_after_open[i], hosted_cer[i]) for i in range(len(walls))]
                remaining = max(0.0, target_cer - sum(cer_alloc))
                caps = [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
                dist_cer, _ = allocate_capped(remaining, lengths, caps)
                cer_alloc = [cer_alloc[i] + dist_cer[i] for i in range(len(walls))]

            # Plaster allocation (depends on project setting)
            plaster_caps = net_after_open if plaster_under else [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
            target_plaster = min(ssot_plaster, sum(plaster_caps)) if ssot_plaster > 0 else 0.0
            plaster_alloc, _ = allocate_capped(target_plaster, plaster_caps, plaster_caps)

            for iwn, wn in enumerate(names):
                wl = lengths[iwn]
                whh = heights[iwn]
                gross = gross_list[iwn]
                note = ''
                if notes_open:
                    note = notes_open
                if notes_cer and not plaster_under:
                    note = (note + '؛ ' if note else '') + notes_cer
                if isinstance(walls[iwn], dict) and walls[iwn].get('_pseudo'):
                    note = (note + '؛ ' if note else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'
                cer_deduct = 0.0 if plaster_under else cer_alloc[iwn]
                write_row(ws2, row, [
                    seq, rname, wn, wl, whh,
                    gross, open_alloc[iwn], cer_deduct, plaster_alloc[iwn], note
                ])
                row += 1
                seq += 1

            # Ceiling row from SSOT
            if d and fnum(d.plaster_ceiling) > 0:
                write_row(ws2, row, [seq, rname, 'سقف', '-', '-', fnum(d.plaster_ceiling), 0.0, 0.0, fnum(d.plaster_ceiling), ''])
                row += 1
                seq += 1
        auto_fit(ws2)

    # --- SHEET: PAINT ---
    if ensure_sheet('paint'):
        ws = wb.create_sheet("الدهان")
        setup_sheet(ws, "الدهان")
        write_header(ws, ["م", "الغرفة", "دهان جدران", "دهان سقف", "الإجمالي"])
        
        for i, r in enumerate(project.rooms, 2):
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            if d:
                write_row(ws, i, [i-1, name, d.paint_walls, d.paint_ceiling, d.paint_total])

        auto_fit(ws)

        # Detailed per-wall paint book (standalone)
        ws2 = wb.create_sheet("الدهان (مفصل)")
        setup_sheet(ws2, "الدهان (مفصل)")
        write_header(ws2, [
            "م", "الغرفة", "الجدار/السقف", "الطول (م)", "الارتفاع (م)",
            "صافي بعد الفتحات (م²)", "سيراميك (م²)", "صافي الدهان (م²)", "ملاحظات"
        ])
        row = 2
        seq = 1
        zones = getattr(project, 'ceramic_zones', []) or []

        for r in getattr(project, 'rooms', []) or []:
            rname = str(val(r, 'name', '') or '').strip()
            walls = iter_room_walls(r)
            if not walls:
                continue
            d = rooms_map.get(rname)
            ssot_open = fnum(getattr(d, 'walls_openings', 0.0) or 0.0) if d else 0.0
            ssot_cer = fnum(getattr(d, 'ceramic_wall', 0.0) or 0.0) if d else 0.0
            ssot_paint = fnum(getattr(d, 'paint_walls', 0.0) or 0.0) if d else 0.0

            openings = room_openings_map.get(rname, [])
            hosted = [o for o in openings if val(o, 'host_wall', None)]

            lengths = [wall_length(w) for w in walls]
            heights = [wall_height(w, r) for w in walls]
            names = [wall_name(w, i) for i, w in enumerate(walls)]
            gross_list = [max(0.0, lengths[i] * heights[i]) for i in range(len(walls))]

            hosted_dem = [0.0 for _ in walls]
            for o in hosted:
                host = norm_text(val(o, 'host_wall', ''))
                if not host:
                    continue
                for iwn, wn in enumerate(names):
                    if host == norm_text(wn):
                        hosted_dem[iwn] += opening_area_for_room(o, rname)
                        break
            total_hosted = sum(hosted_dem)
            if ssot_open <= 0 and total_hosted > 0:
                ssot_open = total_hosted
            scale = 1.0
            notes_open = ''
            if total_hosted > 0 and ssot_open > 0 and total_hosted > ssot_open:
                scale = ssot_open / total_hosted
                notes_open = 'تم تحجيم فتحات مستضافة لتطابق SSOT'
            hosted_alloc = [min(gross_list[i], hosted_dem[i] * scale) for i in range(len(walls))]
            rem_caps = [max(0.0, gross_list[i] - hosted_alloc[i]) for i in range(len(walls))]
            rem_open = max(0.0, ssot_open - sum(hosted_alloc))
            dist_open, unalloc_open = allocate_capped(rem_open, lengths, rem_caps)
            open_alloc = [hosted_alloc[i] + dist_open[i] for i in range(len(walls))]
            if unalloc_open > 0.01:
                notes_open = (notes_open + '؛ ' if notes_open else '') + 'فتحـات أكبر من طاقة الجدران'
            net_after_open = [max(0.0, gross_list[i] - open_alloc[i]) for i in range(len(walls))]

            # Ceramic allocation matched to SSOT
            room_z = [z for z in zones if norm_text(val(z, 'room_name', '')) == norm_text(rname)]
            wall_z = [z for z in room_z if norm_text(val(z, 'surface_type', 'wall')) == 'wall']
            z_hosted = [z for z in wall_z if val(z, 'wall_name', None)]
            hosted_cer = [0.0 for _ in walls]
            for z in z_hosted:
                wname = norm_text(val(z, 'wall_name', ''))
                if not wname:
                    continue
                for iwn, wn in enumerate(names):
                    if wname == norm_text(wn):
                        hosted_cer[iwn] += fnum(safe_zone_area(z))
                        break
            target_cer = min(ssot_cer, sum(net_after_open)) if ssot_cer > 0 else 0.0
            sum_hosted_cer = sum(hosted_cer)
            cer_alloc = [0.0 for _ in walls]
            notes_cer = ''
            if sum_hosted_cer > 0 and target_cer > 0 and sum_hosted_cer > target_cer:
                s = target_cer / sum_hosted_cer
                cer_alloc = [min(net_after_open[i], hosted_cer[i] * s) for i in range(len(walls))]
                notes_cer = 'تم تحجيم سيراميك مستضاف لتطابق SSOT'
            else:
                cer_alloc = [min(net_after_open[i], hosted_cer[i]) for i in range(len(walls))]
                remaining = max(0.0, target_cer - sum(cer_alloc))
                caps = [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
                dist_cer, _ = allocate_capped(remaining, lengths, caps)
                cer_alloc = [cer_alloc[i] + dist_cer[i] for i in range(len(walls))]

            # Paint allocation matched to SSOT
            paint_caps = [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
            target_paint = min(ssot_paint, sum(paint_caps)) if ssot_paint > 0 else 0.0
            paint_alloc, _ = allocate_capped(target_paint, paint_caps, paint_caps)

            for iwn, wn in enumerate(names):
                wl = lengths[iwn]
                whh = heights[iwn]
                notes = notes_open
                if notes_cer:
                    notes = (notes + '؛ ' if notes else '') + notes_cer
                if isinstance(walls[iwn], dict) and walls[iwn].get('_pseudo'):
                    notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'
                write_row(ws2, row, [seq, rname, wn, wl, whh, net_after_open[iwn], cer_alloc[iwn], paint_alloc[iwn], notes])
                row += 1
                seq += 1

            if d and (fnum(d.paint_ceiling) > 0 or fnum(d.ceramic_ceiling) > 0):
                base = fnum(d.paint_ceiling) + min(fnum(d.paint_ceiling), fnum(d.ceramic_ceiling))
                write_row(ws2, row, [seq, rname, 'سقف', '-', '-', base, fnum(d.ceramic_ceiling), fnum(d.paint_ceiling), ''])
                row += 1
                seq += 1
        auto_fit(ws2)

    # --- SHEET: CERAMIC ---
    if ensure_sheet('ceramic'):
        ws = wb.create_sheet("السيراميك")
        setup_sheet(ws, "السيراميك")

        # Per-room SSOT totals
        write_header(ws, ["م", "الغرفة", "سيراميك جدران", "سيراميك أرضيات", "سيراميك سقف", "الإجمالي"])
        row = 2
        for i, r in enumerate(project.rooms, 2):
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            if d:
                write_row(ws, row, [row-1, name, d.ceramic_wall, d.ceramic_floor, d.ceramic_ceiling, d.ceramic_wall + d.ceramic_floor + d.ceramic_ceiling])
            else:
                write_row(ws, row, [row-1, name, 0, 0, 0, 0])
            row += 1

        # Zones details table
        row += 2
        ws.cell(row=row, column=2, value="تفاصيل مناطق السيراميك (Zones)").font = FONT_BOLD
        row += 1
        write_header(ws, [
            "م", "اسم المنطقة", "الغرفة", "الجدار", "النوع", "المحيط (م)", "الارتفاع (م)", "بداية (م)",
            "المساحة (م²)", "Effective Area", "لاصق (كغ)", "جراوت (كغ)"
        ])
        row += 1
        zones = getattr(project, 'ceramic_zones', []) or []
        for idx, z in enumerate(zones, 1):
            z_name = val(z, 'name', '')
            z_room = val(z, 'room_name', '')
            z_wall = val(z, 'wall_name', '')
            stype = val(z, 'surface_type', 'wall')
            perim = fnum(val(z, 'perimeter', 0.0))
            h = fnum(val(z, 'height', 0.0))
            start_h = fnum(val(z, 'start_height', 0.0))
            eff_area = fnum(val(z, 'effective_area', 0.0))
            area = fnum(safe_zone_area(z))
            adh = fnum(val(z, 'adhesive_kg', 0.0))
            grout = fnum(val(z, 'grout_kg', 0.0))
            write_row(ws, row, [idx, z_name, z_room, z_wall, stype, perim, h, start_h, area, eff_area, adh, grout])
            row += 1

        auto_fit(ws)

        # Detailed per-wall ceramic book (standalone)
        ws2 = wb.create_sheet("السيراميك (مفصل)")
        setup_sheet(ws2, "السيراميك (مفصل)")
        write_header(ws2, [
            "م", "نوع السطر", "الغرفة", "الجدار/الأرضية/السقف", "الطول (م)", "الارتفاع (م)", "السيراميك (م²)", "ملاحظات"
        ])
        row2 = 2
        seq2 = 1
        zones = getattr(project, 'ceramic_zones', []) or []
        for r in getattr(project, 'rooms', []) or []:
            rname = str(val(r, 'name', '') or '').strip()
            walls = iter_room_walls(r)
            if not walls:
                continue
            d = rooms_map.get(rname)
            ssot_wall = fnum(getattr(d, 'ceramic_wall', 0.0) or 0.0) if d else 0.0
            ssot_floor = fnum(getattr(d, 'ceramic_floor', 0.0) or 0.0) if d else 0.0
            ssot_ceil = fnum(getattr(d, 'ceramic_ceiling', 0.0) or 0.0) if d else 0.0

            lengths = [wall_length(w) for w in walls]
            heights = [wall_height(w, r) for w in walls]
            names = [wall_name(w, i) for i, w in enumerate(walls)]
            gross_list = [max(0.0, lengths[i] * heights[i]) for i in range(len(walls))]
            # Ceramic should never exceed wall gross; use gross as capacity here.
            caps_wall = gross_list

            room_z = [z for z in zones if norm_text(val(z, 'room_name', '')) == norm_text(rname)]
            wall_z = [z for z in room_z if norm_text(val(z, 'surface_type', 'wall')) == 'wall']
            floor_z = [z for z in room_z if norm_text(val(z, 'surface_type', '')) == 'floor']
            ceil_z = [z for z in room_z if norm_text(val(z, 'surface_type', '')) == 'ceiling']
            z_hosted = [z for z in wall_z if val(z, 'wall_name', None)]

            hosted_cer = [0.0 for _ in walls]
            for z in z_hosted:
                wname = norm_text(val(z, 'wall_name', ''))
                if not wname:
                    continue
                for iwn, wn in enumerate(names):
                    if wname == norm_text(wn):
                        hosted_cer[iwn] += fnum(safe_zone_area(z))
                        break

            target = min(ssot_wall, sum(caps_wall)) if ssot_wall > 0 else 0.0
            sum_hosted = sum(hosted_cer)
            notes_common = ''
            cer_alloc = [0.0 for _ in walls]
            if sum_hosted > 0 and target > 0 and sum_hosted > target:
                s = target / sum_hosted
                cer_alloc = [min(caps_wall[i], hosted_cer[i] * s) for i in range(len(walls))]
                notes_common = 'تم تحجيم سيراميك مستضاف لتطابق SSOT'
            else:
                cer_alloc = [min(caps_wall[i], hosted_cer[i]) for i in range(len(walls))]
                remaining = max(0.0, target - sum(cer_alloc))
                caps = [max(0.0, caps_wall[i] - cer_alloc[i]) for i in range(len(walls))]
                dist, _ = allocate_capped(remaining, lengths, caps)
                cer_alloc = [cer_alloc[i] + dist[i] for i in range(len(walls))]

            for iwn, wn in enumerate(names):
                notes = notes_common
                if isinstance(walls[iwn], dict) and walls[iwn].get('_pseudo'):
                    notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'
                write_row(ws2, row2, [seq2, 'جدار', rname, wn, lengths[iwn], heights[iwn], cer_alloc[iwn], notes])
                row2 += 1
                seq2 += 1

            if ssot_floor > 0:
                write_row(ws2, row2, [seq2, 'أرضية', rname, 'أرضية', '-', '-', ssot_floor, 'SSOT'])
                row2 += 1
                seq2 += 1

            if ssot_ceil > 0:
                write_row(ws2, row2, [seq2, 'سقف', rname, 'سقف', '-', '-', ssot_ceil, 'SSOT'])
                row2 += 1
                seq2 += 1

        auto_fit(ws2)

    # --- SHEET: OPENINGS (Doors + Windows) ---
    if ensure_sheet('openings'):
        ws = wb.create_sheet("الفتحات")
        setup_sheet(ws, "الفتحات")
        write_header(ws, [
            "م", "النوع", "الاسم", "المادة", "العرض", "الارتفاع", "العدد",
            "المساحة/قطعة", "المساحة الإجمالية", "منسوب (م)", "الجدار", "الغرف المرتبطة", "كميات الغرف"
        ])

        row = 2
        for o in (getattr(project, 'doors', []) or []) + (getattr(project, 'windows', []) or []):
            otype = str(val(o, 'opening_type', '') or '').upper() or ('DOOR' if o in (getattr(project, 'doors', []) or []) else 'WINDOW')
            name = val(o, 'name', '')
            mat = val(o, 'material_type', None) or val(o, 'type', '')
            w = fnum(val(o, 'width', None) or val(o, 'w', 0.0))
            h = fnum(val(o, 'height', None) or val(o, 'h', 0.0))
            qty = int(val(o, 'quantity', None) or val(o, 'qty', 1) or 1)
            place = fnum(val(o, 'placement_height', 0.0))
            host = val(o, 'host_wall', '')
            assigned = val(o, 'assigned_rooms', None)
            if assigned is None:
                assigned = list((val(o, 'room_quantities', {}) or {}).keys())
            qmap = val(o, 'room_quantities', {}) or {}
            write_row(ws, row, [
                row-1, otype, name, mat, w, h, qty,
                (w * h) if (w > 0 and h > 0) else 0.0,
                get_opening_area(o), place, host,
                ", ".join(assigned) if isinstance(assigned, list) else str(assigned or ''),
                str(qmap)
            ])
            row += 1

        auto_fit(ws)

    # --- SHEET: BASEBOARDS ---
    if ensure_sheet('baseboards'):
        ws = wb.create_sheet("النعلات")
        setup_sheet(ws, "النعلات")

        # SSOT per-room
        write_header(ws, ["م", "الغرفة", "نعلات (م.ط)"])
        row = 2
        for i, r in enumerate(project.rooms, 2):
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            write_row(ws, row, [row-1, name, (d.baseboard_length if d else 0.0)])
            row += 1

        # Items ledger (if user uses Baseboard model)
        row += 2
        ws.cell(row=row, column=2, value="تفاصيل النعلات (عناصر)" ).font = FONT_BOLD
        row += 1
        write_header(ws, ["م", "الاسم", "المحيط", "خصم الأبواب", "الصافي (م.ط)", "النوع", "الارتفاع (سم)", "المساحة (م²)", "لاصق (كغ)"])
        row += 1
        bbs = getattr(project, 'baseboards', []) or []
        for idx, b in enumerate(bbs, 1):
            perim = fnum(val(b, 'perimeter', 0.0))
            ded = fnum(val(b, 'door_width_deduction', 0.0))
            net = fnum(val(b, 'net_length_m', None))
            if net <= 0 and perim > 0:
                net = max(0.0, perim - ded)
            mat = val(b, 'material_type', '')
            hcm = fnum(val(b, 'height_cm', 0.0))
            area = fnum(val(b, 'area_m2', None))
            if area <= 0 and net > 0 and hcm > 0:
                area = net * (hcm / 100.0)
            adh = fnum(val(b, 'adhesive_kg', None))
            write_row(ws, row, [idx, val(b, 'name', ''), perim, ded, net, mat, hcm, area, adh])
            row += 1

        auto_fit(ws)

    # --- SHEET: STONE ---
    if ensure_sheet('stone'):
        ws = wb.create_sheet("الحجر")
        setup_sheet(ws, "الحجر")

        write_header(ws, ["م", "الغرفة", "حجر/أطر (م.ط) - SSOT"])
        row = 2
        for r in project.rooms:
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            write_row(ws, row, [row-1, name, (d.stone_length if d else 0.0)])
            row += 1

        row += 2
        ws.cell(row=row, column=2, value="تفاصيل الحجر حسب الفتحات" ).font = FONT_BOLD
        row += 1
        write_header(ws, ["م", "النوع", "الاسم", "العرض", "الارتفاع", "العدد", "الحجر (م.ط)"])
        row += 1
        idx = 1
        for o in (getattr(project, 'doors', []) or []) + (getattr(project, 'windows', []) or []):
            otype = str(val(o, 'opening_type', '') or '').upper() or ('DOOR' if o in (getattr(project, 'doors', []) or []) else 'WINDOW')
            name = val(o, 'name', '')
            w = fnum(val(o, 'width', None) or val(o, 'w', 0.0))
            h = fnum(val(o, 'height', None) or val(o, 'h', 0.0))
            qty = int(val(o, 'quantity', None) or val(o, 'qty', 1) or 1)
            write_row(ws, row, [idx, otype, name, w, h, qty, get_opening_stone(o)])
            row += 1
            idx += 1

        auto_fit(ws)

    # --- SHEET: WALLS ---
    if ensure_sheet('walls'):
        ws = wb.create_sheet("المباني")
        setup_sheet(ws, "المباني")
        write_header(ws, [
            "م", "اسم الجدار", "الغرفة", "الطبقة", "الطول", "الارتفاع",
            "قائم (م²)", "خصم (م²)", "صافي (م²)", "نوع السطح", "اتجاه", "سيراميك (ارتفاع)", "سيراميك (مساحة)"
        ])

        row = 2
        walls = getattr(project, 'walls', []) or []
        # If global walls list is empty, fall back to per-room walls.
        if walls:
            for idx, w in enumerate(walls, 1):
                write_row(ws, row, [
                    idx,
                    val(w, 'name', ''),
                    val(w, 'room_name', '') or val(w, 'room', ''),
                    val(w, 'layer', ''),
                    fnum(val(w, 'length', 0.0)),
                    fnum(val(w, 'height', 0.0)),
                    fnum(val(w, 'gross_area', None) or val(w, 'gross', 0.0)),
                    fnum(val(w, 'deduction_area', None) or val(w, 'deduct', 0.0)),
                    fnum(val(w, 'net_area', None) or val(w, 'net', 0.0)),
                    val(w, 'surface_type', ''),
                    val(w, 'direction', ''),
                    fnum(val(w, 'ceramic_height', 0.0)),
                    fnum(val(w, 'ceramic_area', 0.0)),
                ])
                row += 1
        else:
            idx = 1
            for r in getattr(project, 'rooms', []) or []:
                rname = val(r, 'name', '')
                for w in (val(r, 'walls', []) or []):
                    length = fnum(val(w, 'length', 0.0))
                    height = fnum(val(w, 'height', 0.0) or val(r, 'wall_height', 0.0) or 3.0)
                    gross = fnum(val(w, 'gross_area', None) or val(w, 'gross', None) or (length * height))
                    deduct = fnum(val(w, 'deduction_area', None) or val(w, 'deduct', 0.0))
                    net = fnum(val(w, 'net_area', None) or val(w, 'net', None) or max(0.0, gross - deduct))
                    write_row(ws, row, [
                        idx,
                        val(w, 'name', ''),
                        rname,
                        val(w, 'layer', ''),
                        length,
                        height,
                        gross,
                        deduct,
                        net,
                        val(w, 'surface_type', ''),
                        val(w, 'direction', ''),
                        fnum(val(w, 'ceramic_height', 0.0)),
                        fnum(val(w, 'ceramic_area', 0.0)),
                    ])
                    row += 1
                    idx += 1

        auto_fit(ws)

    # --- SHEET: CEILING TILES ---
    if ensure_sheet('ceiling_tiles'):
        ws = wb.create_sheet("بلاط الأسقف")
        setup_sheet(ws, "بلاط الأسقف")
        # No dedicated ceiling-tiles ledger exists; export ceiling areas as a takeoff baseline.
        write_header(ws, ["م", "الغرفة", "مساحة السقف (م²)"])
        row = 2
        for r in getattr(project, 'rooms', []) or []:
            name = val(r, 'name', '-')
            d = rooms_map.get(name)
            write_row(ws, row, [row-1, name, (d.ceiling_area if d else fnum(val(r, 'area', 0.0)))])
            row += 1
        auto_fit(ws)

    # --- SHEET: FLOOR TILES ---
    if ensure_sheet('floor_tiles'):
        ws = wb.create_sheet("بلاط الأرضيات")
        setup_sheet(ws, "بلاط الأرضيات")
        write_header(ws, ["م", "الوصف", "الصافي (م²)", "مع الهدر (م²)"])

        items = getattr(project, 'tiles_items', []) or []
        waste_pct = fnum(getattr(project, 'tiles_waste_percentage', 0.0))
        total = 0.0
        row = 2
        for idx, it in enumerate(items, 1):
            desc = val(it, 'desc', None) or val(it, 'description', '')
            area = fnum(val(it, 'area', 0.0))
            with_waste = area * (1.0 + (waste_pct / 100.0))
            write_row(ws, row, [idx, desc, area, with_waste])
            total += area
            row += 1

        row += 1
        ws.cell(row=row, column=2, value="الإجمالي" ).font = FONT_BOLD
        ws.cell(row=row, column=3, value=total).font = FONT_BOLD
        ws.cell(row=row, column=4, value=total * (1.0 + (waste_pct / 100.0))).font = FONT_BOLD
        auto_fit(ws)

    try:
        wb.save(filepath)
        if status_cb: status_cb(f"Saved: {filepath}", "✅")
        return True
    except Exception as e:
        if status_cb: status_cb(f"Error: {e}", "❌")
        return False
