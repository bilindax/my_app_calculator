"""
Master Sheet Export (دفتر المساحة الموحد)
=========================================
Exports a single comprehensive Excel sheet containing all quantities in a centralized view.
Order: Total -> Openings -> Plaster -> Ceramic -> Paint.
"""

from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import Any, Callable, Optional

from ..calculations.helpers import safe_float, safe_zone_area

# ===== STYLES =====
COLOR_HEADER_BG = "1F4E78"  # Dark Blue
COLOR_HEADER_FG = "FFFFFF"
COLOR_TOTAL_BG = "BDD7EE"
COLOR_GROUP_1 = "E2EFDA" # Light Green
COLOR_GROUP_2 = "FFF2CC" # Light Yellow
COLOR_GROUP_3 = "FCE4D6" # Light Orange
COLOR_GROUP_4 = "DDEBF7" # Light Blue

FONT_HEADER = Font(name='Segoe UI', bold=True, size=10, color=COLOR_HEADER_FG)
FONT_NORMAL = Font(name='Segoe UI', size=10)
FONT_BOLD = Font(name='Segoe UI', bold=True, size=10)

BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def setup_sheet(ws, title, is_rtl=True):
    ws.title = title
    if is_rtl:
        ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

def write_header(ws, headers, row=1):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = FONT_HEADER
        cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

def write_row(ws, row_idx, data, colors=None):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = FONT_NORMAL
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER
        
        if isinstance(val, float):
            cell.number_format = '0.00'
            
        if colors and col in colors:
            cell.fill = PatternFill(start_color=colors[col], end_color=colors[col], fill_type="solid")

def auto_fit(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if '\n' in str(cell.value):
                        val_len = max(len(line) for line in str(cell.value).split('\n'))
                    max_length = max(max_length, val_len)
            except:
                pass
        adjusted_width = min(max_length + 4, 40)
        ws.column_dimensions[column_letter].width = adjusted_width


def _norm_text(val: Any) -> str:
    try:
        return str(val or '').strip().lower()
    except Exception:
        return ''


def _get_opening_field(o: Any, dict_key: str, attr: str, default: Any = None) -> Any:
    if isinstance(o, dict):
        return o.get(dict_key, default)
    return getattr(o, attr, default)


def _opening_area_each(o: Any) -> float:
    if isinstance(o, dict):
        if o.get('area_each') is not None:
            return safe_float(o.get('area_each'))
        w = safe_float(o.get('width', o.get('w', 0.0)))
        h = safe_float(o.get('height', o.get('h', 0.0)))
        return max(0.0, w * h)
    try:
        return safe_float(getattr(o, 'area_each', 0.0))
    except Exception:
        w = safe_float(getattr(o, 'width', 0.0))
        h = safe_float(getattr(o, 'height', 0.0))
        return max(0.0, w * h)


def _opening_total_area(o: Any) -> float:
    if isinstance(o, dict):
        if o.get('area') is not None:
            return safe_float(o.get('area'))
        return _opening_area_each(o) * max(1, int(o.get('qty', o.get('quantity', 1)) or 1))
    try:
        return safe_float(getattr(o, 'area', 0.0))
    except Exception:
        qty = max(1, int(getattr(o, 'quantity', 1) or 1))
        return _opening_area_each(o) * qty


def _opening_type_label(o: Any) -> str:
    ot = _get_opening_field(o, 'opening_type', 'opening_type', None)
    if not ot and isinstance(o, dict):
        # Legacy heuristic
        ot = 'WINDOW' if safe_float(o.get('glass', 0.0)) > 0 else 'DOOR'
    ot = str(ot or '').upper()
    return 'شباك' if ot == 'WINDOW' else 'باب'


def _rooms_for_opening(opening: Any, rooms: list) -> list:
    """Best-effort list of rooms this opening belongs to."""
    name = _get_opening_field(opening, 'name', 'name', '')
    assigned_rooms = _get_opening_field(opening, 'assigned_rooms', 'assigned_rooms', None) or []
    result = []
    seen = set()
    for rname in assigned_rooms:
        rn = str(rname or '').strip()
        if rn and rn not in seen:
            seen.add(rn)
            result.append(rn)
    # Also check room.opening_ids
    for r in rooms:
        rname = (getattr(r, 'name', None) if not isinstance(r, dict) else r.get('name')) or ''
        ids = (getattr(r, 'opening_ids', None) if not isinstance(r, dict) else r.get('opening_ids')) or []
        if name and name in ids:
            rn = str(rname or '').strip()
            if rn and rn not in seen:
                seen.add(rn)
                result.append(rn)
    return result


def _room_opening_area(opening: Any, room_name: str, rooms_for_opening: list) -> float:
    """Return opening area attributed to room_name based on shares/room_quantities.

    This is export-only attribution (does not affect SSOT totals).
    """
    room_name = str(room_name or '').strip()
    if not room_name:
        return 0.0

    room_shares = _get_opening_field(opening, 'room_shares', 'room_shares', None) or {}
    if room_shares:
        share = safe_float(room_shares.get(room_name, 0.0))
        return _opening_total_area(opening) * max(0.0, share)

    assigned_rooms = _get_opening_field(opening, 'assigned_rooms', 'assigned_rooms', None) or []
    share_mode = _get_opening_field(opening, 'share_mode', 'share_mode', None)
    if assigned_rooms:
        if room_name not in assigned_rooms:
            return 0.0
        if (share_mode or 'split') == 'single':
            return _opening_total_area(opening) if room_name == str(assigned_rooms[0] or '').strip() else 0.0
        # default split/custom: equal
        return _opening_total_area(opening) / max(1, len(assigned_rooms))

    # Fallback: if referenced by room.opening_ids and appears in multiple rooms -> equal split
    if rooms_for_opening:
        if room_name not in rooms_for_opening:
            return 0.0
        return _opening_total_area(opening) / max(1, len(rooms_for_opening))

    return 0.0


def _room_opening_area_with_qty(opening: Any, room_name: str, rooms_for_opening: list) -> float:
    """Like _room_opening_area, but honors room_quantities if present."""
    rq = _get_opening_field(opening, 'room_quantities', 'room_quantities', None) or {}
    if rq and room_name in rq:
        room_qty = max(0, int(rq.get(room_name, 0) or 0))
        return _opening_area_each(opening) * room_qty
    return _room_opening_area(opening, room_name, rooms_for_opening)


def _iter_room_walls(room_obj: Any, project_default_h: float) -> list:
    walls = getattr(room_obj, 'walls', None) or []
    if walls:
        return list(walls)
    # Fallback: no explicit walls -> pseudo wall using perimeter
    perim = safe_float(getattr(room_obj, 'perimeter', 0.0) or 0.0)
    h = safe_float(getattr(room_obj, 'wall_height', 0.0) or 0.0) or project_default_h
    if perim > 0 and h > 0:
        return [
            {
                'name': 'محيط الغرفة',
                'length': perim,
                'height': h,
                '_pseudo': True,
            }
        ]
    return []


def _wall_name(w: Any, idx: int) -> str:
    name = (w.get('name') if isinstance(w, dict) else getattr(w, 'name', None))
    name = str(name or '').strip()
    return name if name else f"Wall {idx + 1}"


def _wall_length(w: Any) -> float:
    return safe_float(w.get('length') if isinstance(w, dict) else getattr(w, 'length', 0.0))


def _wall_height(w: Any, room_obj: Any, project_default_h: float) -> float:
    wh = safe_float(w.get('height') if isinstance(w, dict) else getattr(w, 'height', 0.0))
    if wh > 0:
        return wh
    rh = safe_float(getattr(room_obj, 'wall_height', 0.0) or 0.0)
    return rh if rh > 0 else project_default_h


def _allocate_capped(total: float, weights: list, caps: list) -> tuple[list, float]:
    """Allocate `total` proportionally by `weights`, capped by `caps` per item.

    Returns (allocations, unallocated).
    """
    total = float(total or 0.0)
    caps_left = [max(0.0, float(c or 0.0)) for c in caps]
    wts = [max(0.0, float(w or 0.0)) for w in weights]
    alloc = [0.0 for _ in caps_left]
    remaining = total
    active = [i for i, c in enumerate(caps_left) if c > 0]
    eps = 1e-9

    while remaining > eps and active:
        wsum = sum(wts[i] for i in active)
        if wsum <= eps:
            wsum = float(len(active))
            for i in active:
                wts[i] = 1.0

        exhausted = []
        for i in active:
            share = remaining * (wts[i] / wsum) if wsum > eps else 0.0
            if share <= caps_left[i] + eps:
                alloc[i] += share
                caps_left[i] -= share
            else:
                alloc[i] += caps_left[i]
                remaining -= caps_left[i]
                caps_left[i] = 0.0
                exhausted.append(i)

        if not exhausted:
            remaining = 0.0
            break
        active = [i for i in active if caps_left[i] > eps]

    return alloc, max(0.0, remaining)


def _compute_room_wall_breakdown(
    project: Any,
    room_obj: Any,
    rc: Any,
    room_openings: list,
    project_default_h: float,
) -> dict:
    walls = _iter_room_walls(room_obj, project_default_h)
    names = [_wall_name(w, i) for i, w in enumerate(walls)]
    lengths = [max(0.0, _wall_length(w)) for w in walls]
    heights = [max(0.0, _wall_height(w, room_obj, project_default_h)) for w in walls]
    gross_geom = [lengths[i] * heights[i] for i in range(len(walls))]

    notes = []

    # Scale gross areas to match SSOT walls_gross if available
    gross = list(gross_geom)
    if rc is not None:
        ssot_gross = safe_float(getattr(rc, 'walls_gross', 0.0) or 0.0)
        geom_sum = sum(gross_geom)
        if ssot_gross > 0 and geom_sum > 0 and abs(ssot_gross - geom_sum) > 0.01:
            scale = ssot_gross / geom_sum
            gross = [g * scale for g in gross_geom]
            notes.append('تم تحجيم مساحة الجدران لمطابقة SSOT')

    # Opening allocations (SSOT walls_openings)
    rooms_list = getattr(project, 'rooms', []) or []
    per_open_rooms = {}
    for o in room_openings:
        oname = _get_opening_field(o, 'name', 'name', None)
        if oname:
            per_open_rooms[oname] = _rooms_for_opening(o, rooms_list)

    hosted_dem = [0.0 for _ in walls]
    for o in room_openings:
        host = _norm_text(_get_opening_field(o, 'host_wall', 'host_wall', ''))
        if not host:
            continue
        oname = _get_opening_field(o, 'name', 'name', '')
        a = _room_opening_area_with_qty(o, getattr(room_obj, 'name', ''), per_open_rooms.get(oname, []))
        for wi, wn in enumerate(names):
            if host == _norm_text(wn):
                hosted_dem[wi] += a
                break

    total_hosted = sum(hosted_dem)
    ssot_open = safe_float(getattr(rc, 'walls_openings', 0.0) if rc is not None else 0.0)
    if ssot_open <= 0 and total_hosted > 0:
        ssot_open = total_hosted

    scale_open = 1.0
    if total_hosted > 0 and ssot_open > 0 and total_hosted > ssot_open:
        scale_open = ssot_open / total_hosted
        notes.append('تم تحجيم فتحات مستضافة لتطابق SSOT')

    hosted_alloc = [min(gross[i], hosted_dem[i] * scale_open) for i in range(len(walls))]
    rem_caps = [max(0.0, gross[i] - hosted_alloc[i]) for i in range(len(walls))]
    rem_open = max(0.0, ssot_open - sum(hosted_alloc))
    dist_open, unalloc_open = _allocate_capped(rem_open, lengths, rem_caps)
    open_alloc = [hosted_alloc[i] + dist_open[i] for i in range(len(walls))]
    if unalloc_open > 0.01:
        notes.append('فتحـات أكبر من طاقة الجدران')

    net_after_open = [max(0.0, gross[i] - open_alloc[i]) for i in range(len(walls))]

    # Ceramic wall allocations (SSOT ceramic_wall)
    room_name = getattr(room_obj, 'name', '')
    room_zones = [
        z for z in (getattr(project, 'ceramic_zones', []) or [])
        if _norm_text(getattr(z, 'room_name', None) if not isinstance(z, dict) else z.get('room_name')) == _norm_text(room_name)
    ]
    wall_zones = [
        z for z in room_zones
        if _norm_text((getattr(z, 'surface_type', None) if not isinstance(z, dict) else z.get('surface_type')) or 'wall') == 'wall'
    ]
    hosted_cer = [0.0 for _ in walls]
    for z in wall_zones:
        zw = _norm_text((getattr(z, 'wall_name', None) if not isinstance(z, dict) else z.get('wall_name')) or '')
        if not zw:
            continue
        for wi, wn in enumerate(names):
            if zw == _norm_text(wn):
                hosted_cer[wi] += safe_float(safe_zone_area(z) or 0.0)
                break

    ssot_cer = safe_float(getattr(rc, 'ceramic_wall', 0.0) if rc is not None else 0.0)
    target_cer = min(ssot_cer, sum(net_after_open)) if ssot_cer > 0 else 0.0
    sum_hosted_cer = sum(hosted_cer)
    cer_alloc = [0.0 for _ in walls]

    if sum_hosted_cer > 0 and target_cer > 0 and sum_hosted_cer > target_cer:
        s = target_cer / sum_hosted_cer
        cer_alloc = [min(net_after_open[i], hosted_cer[i] * s) for i in range(len(walls))]
        notes.append('تم تحجيم سيراميك مستضاف لتطابق SSOT')
    else:
        cer_alloc = [min(net_after_open[i], hosted_cer[i]) for i in range(len(walls))]
        remaining = max(0.0, target_cer - sum(cer_alloc))
        caps = [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
        dist_cer, _ = _allocate_capped(remaining, lengths, caps)
        cer_alloc = [cer_alloc[i] + dist_cer[i] for i in range(len(walls))]

    # Paint wall allocations (SSOT paint_walls) with ceramic deduction
    paint_caps = [max(0.0, net_after_open[i] - cer_alloc[i]) for i in range(len(walls))]
    ssot_paint = safe_float(getattr(rc, 'paint_walls', 0.0) if rc is not None else 0.0)
    target_paint = min(ssot_paint, sum(paint_caps)) if ssot_paint > 0 else 0.0
    paint_alloc, _ = _allocate_capped(target_paint, paint_caps, paint_caps)

    # Plaster wall allocations (SSOT plaster_walls), behind ceramic
    ssot_plaster = safe_float(getattr(rc, 'plaster_walls', 0.0) if rc is not None else 0.0)
    target_plaster = min(ssot_plaster, sum(net_after_open)) if ssot_plaster > 0 else 0.0
    plaster_alloc, _ = _allocate_capped(target_plaster, net_after_open, net_after_open)

    return {
        'walls': walls,
        'names': names,
        'lengths': lengths,
        'heights': heights,
        'gross': gross,
        'open_alloc': open_alloc,
        'net_after_open': net_after_open,
        'cer_alloc': cer_alloc,
        'paint_alloc': paint_alloc,
        'plaster_alloc': plaster_alloc,
        'notes': '؛ '.join([n for n in notes if n]),
        'room_zones': room_zones,
    }

def export_master_sheet(
    project: Any,
    filepath: str,
    app: Any = None,
    status_cb: Optional[Callable[[str, str], None]] = None
) -> bool:
    
    if status_cb:
        status_cb("Generating Master Sheet...", "📊")

    wb = openpyxl.Workbook()
    ws = wb.active
    setup_sheet(ws, "دفتر المساحة الموحد")
    
    # Title
    ws['B1'] = f"مشروع: {getattr(project, 'project_name', 'Untitled')}"
    ws['B1'].font = Font(name='Segoe UI', size=14, bold=True)
    ws['B2'] = f"دفتر المساحة الشامل - {datetime.now().strftime('%Y-%m-%d')}"
    
    # Headers
    headers = [
        "م", "اسم الغرفة", "نوع الغرفة", 
        "الطول", "العرض", "الارتفاع", "المحيط", "مساحة الأرضية", # Dimensions
        "مساحة الجدران القائمة", "مساحة السقف", "الإجمالي القائم", # Gross
        "خصم الفتحات", # Openings
        "زريقة جدران (صافي)", "زريقة سقف", "إجمالي الزريقة", # Plaster
        "سيراميك جدران", "سيراميك أرضيات", "سيراميك سقف", "إجمالي السيراميك", # Ceramic
        "دهان جدران", "دهان سقف", "إجمالي الدهان", # Paint
        "ملاحظات"
    ]
    
    # Column Group Colors
    col_colors = {}
    # Dimensions: 4-8 (White/Default)
    # Gross: 9-11 (Group 1)
    for c in range(9, 12): col_colors[c] = COLOR_GROUP_1
    # Openings: 12 (Group 2)
    col_colors[12] = COLOR_GROUP_2
    # Plaster: 13-15 (Group 3)
    for c in range(13, 16): col_colors[c] = COLOR_GROUP_3
    # Ceramic: 16-19 (Group 4)
    for c in range(16, 20): col_colors[c] = COLOR_GROUP_4
    # Paint: 20-22 (Group 1 again or White)
    for c in range(20, 23): col_colors[c] = COLOR_GROUP_1
    
    write_header(ws, headers, row=4)
    
    # Use UnifiedCalculator
    from ..calculations.unified_calculator import UnifiedCalculator
    calc = UnifiedCalculator(project)
    all_rooms = calc.calculate_all_rooms()

    rcalc_by_name = {rc.room_name: rc for rc in all_rooms}

    project_default_h = safe_float(getattr(project, 'default_wall_height', 3.0) or 3.0)
    
    r_idx = 5
    
    # Totals
    t_floor = 0.0
    t_gross_wall = 0.0
    t_ceil = 0.0
    t_gross_total = 0.0
    t_openings = 0.0
    t_plaster_wall = 0.0
    t_plaster_ceil = 0.0
    t_plaster_total = 0.0
    t_cer_wall = 0.0
    t_cer_floor = 0.0
    t_cer_ceil = 0.0
    t_cer_total = 0.0
    t_paint_wall = 0.0
    t_paint_ceil = 0.0
    t_paint_total = 0.0
    
    room_obj_by_name = {getattr(r, 'name', ''): r for r in getattr(project, 'rooms', []) or []}

    for i, r_calc in enumerate(all_rooms, 1):
        # Find original room object for dimensions
        room_obj = room_obj_by_name.get(r_calc.room_name)
        
        w = float(getattr(room_obj, 'width', 0.0) or 0.0)
        l = float(getattr(room_obj, 'length', 0.0) or 0.0)
        h = float(getattr(room_obj, 'wall_height', 0.0) or 0.0)
        perim = float(getattr(room_obj, 'perimeter', 0.0) or 0.0)
        if not perim and room_obj:
             # Fallback perim
             perim = sum(float(getattr(wl, 'length', 0.0) or 0.0) for wl in (getattr(room_obj, 'walls', []) or []))
        
        room_type = getattr(room_obj, 'room_type', '-') if room_obj else '-'
        
        # Dimensions
        dim_l = l if l > 0 else "-"
        dim_w = w if w > 0 else "-"
        # Wall height: check if room has multiple walls with different heights
        if room_obj:
            walls = _iter_room_walls(room_obj, project_default_h)
            if len(walls) > 1:
                heights = set()
                for w_item in walls:
                    wh = _wall_height(w_item, room_obj, project_default_h)
                    if wh > 0:
                        heights.add(round(wh, 2))
                if len(heights) > 1:
                    min_h = min(heights)
                    max_h = max(heights)
                    dim_h = f"{min_h:.2f}-{max_h:.2f}"  # Show range
                elif heights:
                    dim_h = next(iter(heights))
                else:
                    dim_h = h if h > 0 else "-"
            else:
                dim_h = h if h > 0 else "-"
        else:
            dim_h = h if h > 0 else "-"
        
        # Gross
        gross_wall = r_calc.walls_gross
        ceil_area = r_calc.ceiling_area
        gross_total = gross_wall + ceil_area
        
        # Openings
        openings = r_calc.walls_openings
        
        # Plaster
        plaster_wall = r_calc.plaster_walls
        plaster_ceil = r_calc.plaster_ceiling
        plaster_total = r_calc.plaster_total
        
        # Ceramic
        cer_wall = r_calc.ceramic_wall
        cer_floor = r_calc.ceramic_floor
        cer_ceil = r_calc.ceramic_ceiling
        cer_total = cer_wall + cer_floor + cer_ceil
        
        # Paint
        paint_wall = r_calc.paint_walls
        paint_ceil = r_calc.paint_ceiling
        paint_total = r_calc.paint_total
        
        # Accumulate Totals
        t_floor += ceil_area # Floor area is same as ceiling area usually
        t_gross_wall += gross_wall
        t_ceil += ceil_area
        t_gross_total += gross_total
        t_openings += openings
        t_plaster_wall += plaster_wall
        t_plaster_ceil += plaster_ceil
        t_plaster_total += plaster_total
        t_cer_wall += cer_wall
        t_cer_floor += cer_floor
        t_cer_ceil += cer_ceil
        t_cer_total += cer_total
        t_paint_wall += paint_wall
        t_paint_ceil += paint_ceil
        t_paint_total += paint_total
        
        row_data = [
            i, r_calc.room_name, room_type,
            dim_l, dim_w, dim_h, perim, ceil_area,
            gross_wall, ceil_area, gross_total,
            openings,
            plaster_wall, plaster_ceil, plaster_total,
            cer_wall, cer_floor, cer_ceil, cer_total,
            paint_wall, paint_ceil, paint_total,
            ""
        ]
        
        write_row(ws, r_idx, row_data, colors=col_colors)
        r_idx += 1
        
    # Totals Row
    ws.cell(row=r_idx, column=2, value="الإجمالي الكلي").font = FONT_BOLD
    
    totals_map = {
        8: t_floor,
        9: t_gross_wall, 10: t_ceil, 11: t_gross_total,
        12: t_openings,
        13: t_plaster_wall, 14: t_plaster_ceil, 15: t_plaster_total,
        16: t_cer_wall, 17: t_cer_floor, 18: t_cer_ceil, 19: t_cer_total,
        20: t_paint_wall, 21: t_paint_ceil, 22: t_paint_total
    }
    
    for col, val in totals_map.items():
        cell = ws.cell(row=r_idx, column=col, value=val)
        cell.font = FONT_BOLD
        cell.number_format = '0.00'
        cell.fill = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type="solid")
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_CENTER

    auto_fit(ws)

    # =============================
    # Detailed Sheets (Standalone)
    # =============================
    all_openings = (getattr(project, 'doors', []) or []) + (getattr(project, 'windows', []) or [])
    openings_by_name = {}
    for o in all_openings:
        name = _get_opening_field(o, 'name', 'name', None)
        if name:
            openings_by_name[str(name).strip()] = o

    # Precompute room-wise opening lists
    room_opening_ids = {}
    room_opening_objs = {}
    for r in getattr(project, 'rooms', []) or []:
        rname = getattr(r, 'name', '')
        ids = list(getattr(r, 'opening_ids', []) or [])
        room_opening_ids[rname] = ids
        refs = []
        for oid in ids:
            o = openings_by_name.get(str(oid).strip())
            if o is not None:
                refs.append(o)
        # Also add openings that reference this room in assigned_rooms/room_shares/room_quantities
        for o in all_openings:
            assigned = _get_opening_field(o, 'assigned_rooms', 'assigned_rooms', None) or []
            room_shares = _get_opening_field(o, 'room_shares', 'room_shares', None) or {}
            room_qtys = _get_opening_field(o, 'room_quantities', 'room_quantities', None) or {}
            if rname and (rname in assigned or rname in room_shares or rname in room_qtys):
                refs.append(o)
        # De-dup by name
        dedup = {}
        for o in refs:
            nm = _get_opening_field(o, 'name', 'name', None)
            if nm:
                dedup[str(nm).strip()] = o
        room_opening_objs[rname] = list(dedup.values())

    # --- Openings Sheet ---
    ws_open = wb.create_sheet("تفاصيل الفتحات")
    setup_sheet(ws_open, "تفاصيل الفتحات")
    open_headers = [
        "م", "المعرف", "النوع", "الطبقة", "النوع/المادة", "العدد",
        "العرض (م)", "الارتفاع (م)", "منسوب التركيب (م)",
        "مساحة القطعة (م²)", "المساحة الإجمالية (م²)",
        "الجدار المرتبط", "الغرف المرتبطة", "نظام التوزيع", "ملاحظات"
    ]
    write_header(ws_open, open_headers, row=1)
    row_i = 2
    for idx, o in enumerate(all_openings, 1):
        name = _get_opening_field(o, 'name', 'name', '-')
        otype = _opening_type_label(o)
        layer = _get_opening_field(o, 'layer', 'layer', '') or ''
        mat = _get_opening_field(o, 'material_type', 'material_type', _get_opening_field(o, 'type', 'type', '')) or ''
        qty = int(_get_opening_field(o, 'quantity', 'quantity', _get_opening_field(o, 'qty', 'qty', 1)) or 1)
        width = safe_float(_get_opening_field(o, 'width', 'width', _get_opening_field(o, 'w', 'w', 0.0)))
        height = safe_float(_get_opening_field(o, 'height', 'height', _get_opening_field(o, 'h', 'h', 0.0)))
        placement_h = safe_float(_get_opening_field(o, 'placement_height', 'placement_height', 0.0))
        host_wall = _get_opening_field(o, 'host_wall', 'host_wall', '') or ''
        rooms_for = _rooms_for_opening(o, getattr(project, 'rooms', []) or [])
        share_mode = _get_opening_field(o, 'share_mode', 'share_mode', '') or ''
        notes = ''
        if not rooms_for:
            notes = 'غير مرتبط بأي غرفة'
        write_row(ws_open, row_i, [
            idx, name, otype, layer, mat, qty,
            width, height, placement_h,
            _opening_area_each(o), _opening_total_area(o),
            host_wall, ', '.join(rooms_for) if rooms_for else '-', share_mode, notes
        ])
        row_i += 1
    auto_fit(ws_open)

    # --- Walls Details Sheet (with openings + ceramic + paint/plaster cues) ---
    ws_walls = wb.create_sheet("تفاصيل الجدران")
    setup_sheet(ws_walls, "تفاصيل الجدران")
    wall_headers = [
        "م", "اسم الغرفة", "نوع الغرفة", "الجدار", "طول (م)", "ارتفاع (م)",
        "مساحة الجدار (م²)", "فتحات على الجدار (م²)", "صافي بعد الفتحات (م²)",
        "سيراميك على الجدار (م²)", "خصم السيراميك من الدهان (م²)",
        "دهان الجدار (م²)", "ملاحظات"
    ]
    write_header(ws_walls, wall_headers, row=1)
    row_i = 2
    seq = 1
    for r in getattr(project, 'rooms', []) or []:
        rname = getattr(r, 'name', '')
        rtype = getattr(r, 'room_type', '-') or '-'
        rc = rcalc_by_name.get(rname)
        # Wall list (explicit or pseudo)
        walls = _iter_room_walls(r, project_default_h)
        if not walls:
            continue

        room_openings = room_opening_objs.get(rname, [])
        bd = _compute_room_wall_breakdown(project, r, rc, room_openings, project_default_h)
        note_common = bd['notes']

        for w_idx, w in enumerate(bd['walls']):
            wname = bd['names'][w_idx]
            wlen = bd['lengths'][w_idx]
            whei = bd['heights'][w_idx]
            gross = bd['gross'][w_idx]
            open_area = bd['open_alloc'][w_idx]
            net_after_open = bd['net_after_open'][w_idx]
            ceramic_area = bd['cer_alloc'][w_idx]
            paint_area = bd['paint_alloc'][w_idx]
            paint_deduct = max(0.0, min(net_after_open, ceramic_area))

            notes = note_common or ''
            if isinstance(w, dict) and w.get('_pseudo'):
                notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'

            write_row(ws_walls, row_i, [
                seq, rname, rtype, wname, wlen, whei,
                gross, open_area, net_after_open,
                ceramic_area, paint_deduct, paint_area,
                notes
            ])
            row_i += 1
            seq += 1
    auto_fit(ws_walls)

    # --- Plaster Detailed Book ---
    ws_plaster = wb.create_sheet("دفتر الزريقة (مفصل)")
    setup_sheet(ws_plaster, "دفتر الزريقة (مفصل)")
    plaster_headers = [
        "م", "نوع السطر", "اسم الغرفة", "الجدار/السقف", "طول (م)", "ارتفاع (م)",
        "مساحة (م²)", "خصم الفتحات (م²)", "صافي الزريقة (م²)", "ملاحظات"
    ]
    write_header(ws_plaster, plaster_headers, row=1)
    row_i = 2
    seq = 1
    for r in getattr(project, 'rooms', []) or []:
        rname = getattr(r, 'name', '')
        rc = rcalc_by_name.get(rname)
        walls = _iter_room_walls(r, project_default_h)
        if not walls:
            continue
        room_openings = room_opening_objs.get(rname, [])
        bd = _compute_room_wall_breakdown(project, r, rc, room_openings, project_default_h)
        note_common = bd['notes']

        for w_idx, w in enumerate(bd['walls']):
            wname = bd['names'][w_idx]
            wlen = bd['lengths'][w_idx]
            whei = bd['heights'][w_idx]
            gross = bd['gross'][w_idx]
            open_area = bd['open_alloc'][w_idx]
            net_plaster = bd['plaster_alloc'][w_idx]
            notes = note_common or ''
            if isinstance(w, dict) and w.get('_pseudo'):
                notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'
            write_row(ws_plaster, row_i, [seq, 'جدار', rname, wname, wlen, whei, gross, open_area, net_plaster, notes])
            row_i += 1
            seq += 1

        # Ceiling row (standalone)
        rc = rcalc_by_name.get(rname)
        ceil_area = safe_float(getattr(rc, 'plaster_ceiling', 0.0) if rc else 0.0)
        if ceil_area > 0:
            write_row(ws_plaster, row_i, [seq, 'سقف', rname, 'سقف', '-', '-', ceil_area, 0.0, ceil_area, ''])
            row_i += 1
            seq += 1
    auto_fit(ws_plaster)

    # --- Paint Detailed Book (shows ceramic deduction) ---
    ws_paint = wb.create_sheet("دفتر الدهان (مفصل)")
    setup_sheet(ws_paint, "دفتر الدهان (مفصل)")
    paint_headers = [
        "م", "نوع السطر", "اسم الغرفة", "الجدار/السقف", "طول (م)", "ارتفاع (م)",
        "مساحة بعد الفتحات (م²)", "سيراميك (م²)", "صافي الدهان (م²)", "ملاحظات"
    ]
    write_header(ws_paint, paint_headers, row=1)
    row_i = 2
    seq = 1
    for r in getattr(project, 'rooms', []) or []:
        rname = getattr(r, 'name', '')
        rc = rcalc_by_name.get(rname)
        walls = _iter_room_walls(r, project_default_h)
        if not walls:
            continue

        room_openings = room_opening_objs.get(rname, [])
        bd = _compute_room_wall_breakdown(project, r, rc, room_openings, project_default_h)
        note_common = bd['notes']

        for w_idx, w in enumerate(bd['walls']):
            wname = bd['names'][w_idx]
            wlen = bd['lengths'][w_idx]
            whei = bd['heights'][w_idx]
            net_after_open = bd['net_after_open'][w_idx]
            ceramic_area = bd['cer_alloc'][w_idx]
            paint_area = bd['paint_alloc'][w_idx]
            notes = note_common or ''
            if isinstance(w, dict) and w.get('_pseudo'):
                notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'
            write_row(ws_paint, row_i, [
                seq, 'جدار', rname, wname, wlen, whei,
                net_after_open, ceramic_area, paint_area, notes
            ])
            row_i += 1
            seq += 1

        # Ceiling row (standalone)
        rc = rcalc_by_name.get(rname)
        if rc:
            # paint ceiling is already SSOT (may already include any ceramic-ceiling logic)
            paint_ceil = safe_float(getattr(rc, 'paint_ceiling', 0.0) or 0.0)
            cer_ceil = safe_float(getattr(rc, 'ceramic_ceiling', 0.0) or 0.0)
            if paint_ceil > 0 or cer_ceil > 0:
                base_after_open = paint_ceil + min(paint_ceil, cer_ceil)  # best-effort to show deduction column
                write_row(ws_paint, row_i, [
                    seq, 'سقف', rname, 'سقف', '-', '-',
                    base_after_open, cer_ceil, paint_ceil, ''
                ])
                row_i += 1
                seq += 1
    auto_fit(ws_paint)

    # --- Ceramic Detailed Book (per-wall + floor/ceiling) ---
    ws_cer = wb.create_sheet("دفتر السيراميك (مفصل)")
    setup_sheet(ws_cer, "دفتر السيراميك (مفصل)")
    cer_headers = [
        "م", "نوع السطر", "اسم الغرفة", "الجدار/الأرضية/السقف", "طول (م)", "ارتفاع (م)",
        "سيراميك (م²)", "ملاحظات"
    ]
    write_header(ws_cer, cer_headers, row=1)
    row_i = 2
    seq = 1

    for r in getattr(project, 'rooms', []) or []:
        rname = getattr(r, 'name', '')
        rc = rcalc_by_name.get(rname)
        walls = _iter_room_walls(r, project_default_h)
        if not walls:
            continue

        room_openings = room_opening_objs.get(rname, [])
        bd = _compute_room_wall_breakdown(project, r, rc, room_openings, project_default_h)
        note_common = bd['notes']

        for w_idx, w in enumerate(bd['walls']):
            wname = bd['names'][w_idx]
            wlen = bd['lengths'][w_idx]
            whei = bd['heights'][w_idx]
            cer_area = bd['cer_alloc'][w_idx]
            notes = note_common or ''
            if isinstance(w, dict) and w.get('_pseudo'):
                notes = (notes + '؛ ' if notes else '') + 'لا يوجد تقسيم جدران؛ تم استخدام المحيط'

            write_row(ws_cer, row_i, [seq, 'جدار', rname, wname, wlen, whei, cer_area, notes])
            row_i += 1
            seq += 1

        # Floor/Ceiling from SSOT when available
        floor_area = safe_float(getattr(rc, 'ceramic_floor', 0.0) if rc is not None else 0.0)
        ceil_area = safe_float(getattr(rc, 'ceramic_ceiling', 0.0) if rc is not None else 0.0)
        if floor_area > 0:
            write_row(ws_cer, row_i, [seq, 'أرضية', rname, 'أرضية', '-', '-', floor_area, ''])
            row_i += 1
            seq += 1
        if ceil_area > 0:
            write_row(ws_cer, row_i, [seq, 'سقف', rname, 'سقف', '-', '-', ceil_area, ''])
            row_i += 1
            seq += 1

    auto_fit(ws_cer)
    
    try:
        wb.save(filepath)
        if status_cb:
            status_cb(f"تم حفظ الملف: {filepath}", "✅")
        return True
    except Exception as e:
        if status_cb:
            status_cb(f"خطأ في الحفظ: {e}", "❌")
        return False
